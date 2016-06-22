import requests
from convert import unicode_to_ascii
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.utils import get_column_letter


class Scraper:
    '''
    This class take in 4 variables: frequency, tab, unit, final_year
    frequency=0:quarterly | frequency=1:yearly
    tab=0:BCDKT | tab=1:BCKQHDKD| tab=2:BCLCTTTT| tab=3:BCLCTTGT | tab=4:TMBCTC | tab=7:CSTC
    unit can be 1, 1000 or 1000000 (VND)
    '''

    def __init__(self, frequency, tab, unit, final_year):
        self.frequency = frequency
        self.tab = tab
        self.unit = unit
        self.final_year = final_year
        self.data_table = ['']
        self.URL = 'http://www.bvsc.com.vn/FinancialStatements.aspx'

    def do_work(self, stock_symbol):
        self.request_table(stock_symbol)
        self.write_to_spreadsheet(stock_symbol)
        self.data_table = ['']

    def request_table(self, stock_symbol):
        '''Make a request for the tageted data table then put every value into self.data_table as a list'''
        payload = [('SymbolList', stock_symbol), ('Symbol', stock_symbol),
                   ('Cart_ctl00_webPartManager_wp1629714544_wp1440203596_cbFinanceReport_Callback_Param', self.frequency),
                   ('Cart_ctl00_webPartManager_wp1629714544_wp1440203596_cbFinanceReport_Callback_Param', self.tab),
                   ('Cart_ctl00_webPartManager_wp1629714544_wp1440203596_cbFinanceReport_Callback_Param', self.unit),
                   ('Cart_ctl00_webPartManager_wp1629714544_wp1440203596_cbFinanceReport_Callback_Param', self.final_year)]

        r = requests.post(self.URL, data=payload)

        # Processing Html document
        html_doc = unicode_to_ascii(str(r.text)).replace(',', '')
        html_doc = html_doc.strip('                    ]]></CallbackContent>')
        html_doc = html_doc.strip('<CallbackContent><![CDATA[')

        # Parsing Html document
        soup = BeautifulSoup(html_doc, 'html.parser')
        headers = soup.findAll('font')
        for header in headers:
            self.data_table.append(header.string.strip())
        cells = soup.findAll('td')
        for cell in cells:
            if cell.string is not None:
                self.data_table.append(cell.string.strip())

    def write_to_spreadsheet(self, spreadsheet_name):
        '''Write values from self.data_table to spreadsheet'''
        r = 1
        c = 1
        wb = openpyxl.Workbook()
        sheet = wb.active

        # Write sheet
        for item in self.data_table:
            # Number cells
            if item.lstrip('-').isdigit():
                sheet.cell(row=r, column=c).value = int(item)
                sheet.cell(row=r, column=c).number_format = '0'
                sheet.column_dimensions[get_column_letter(c)].width = len(item) * 1.5
                c += 1
            # Emty cells
            elif item == '':
                c += 1
            # Text cells
            else:
                r += 1
                c = 1
                sheet.cell(row=r, column=c).value = item
                sheet.column_dimensions[get_column_letter(c)].width = len(item) * 2
                c += 1

        wb.save(spreadsheet_name + '.xlsx')
